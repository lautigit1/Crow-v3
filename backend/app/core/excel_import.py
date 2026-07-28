"""Lectura de las facturas en Excel que mandan los proveedores.

El archivo llega tal como lo emite el proveedor: sin normalizar, con títulos
y datos de la empresa arriba, encabezados en una fila cualquiera, y números
escritos como los escribe alguien en Argentina ("1.234,56").

Este módulo hace una sola cosa: convertir ese archivo en filas estructuradas.
No decide qué producto se crea ni toca la base -- de eso se ocupa
`routes/imports.py`. La separación importa porque la fase 6 (extracción de
PDF con IA) va a producir exactamente el mismo `list[FilaCruda]` y reusar
todo lo que sigue sin cambios.
"""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

# Claves internas del mapeo. El valor guardado en `suppliers.column_mapping`
# es {clave interna -> nombre del encabezado en el Excel de ese proveedor}.
CAMPOS = ("sku", "name", "quantity", "unit_cost")

# Encabezados que se intentan reconocer solos la primera vez, para que el
# mapeo llegue pre-completado y el usuario solo confirme. Todo en minúscula y
# sin acentos: la comparación se hace normalizada.
_ALIAS: dict[str, tuple[str, ...]] = {
    "sku": ("codigo", "cod", "sku", "articulo", "art", "referencia", "ref", "codigo articulo"),
    "name": ("descripcion", "detalle", "producto", "articulo", "denominacion", "concepto"),
    "quantity": ("cantidad", "cant", "cant.", "unidades", "qty"),
    "unit_cost": ("precio unitario", "p. unitario", "p unit", "precio", "unitario", "costo", "importe unitario"),
}

# Cuántas filas se miran buscando la fila de encabezados antes de rendirse.
# Las facturas suelen traer logo, razón social y domicilio arriba; 25 cubre
# de sobra ese preámbulo sin recorrer el archivo entero.
_MAX_FILAS_ENCABEZADO = 25


@dataclass
class FilaCruda:
    """Una línea de la factura ya interpretada, antes de resolverla contra el
    catálogo. `error` distinto de None significa que la fila no se pudo leer
    y hay que mostrarla igual para que el usuario la corrija a mano."""

    row_number: int
    sku: str
    name: str
    quantity: int
    unit_cost: Decimal | None
    error: str | None = None


class ExcelInvalido(Exception):
    """El archivo no se pudo abrir o no tiene una estructura reconocible."""


def normalizar(texto: Any) -> str:
    """Minúsculas, sin acentos y sin espacios de más.

    Se usa para comparar encabezados: "Descripción", "DESCRIPCION" y
    "descripcion " tienen que ser lo mismo.
    """
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    for con_acento, sin_acento in zip("áéíóúüñ", "aeiouun", strict=True):
        s = s.replace(con_acento, sin_acento)
    return " ".join(s.split())


def _a_decimal(valor: Any) -> Decimal | None:
    """Convierte una celda a Decimal aceptando el formato local.

    openpyxl ya devuelve números reales cuando la celda es numérica; el
    trabajo está en las que vienen como texto, que es lo habitual cuando el
    proveedor exporta desde su sistema. "1.234,56" es mil doscientos treinta
    y cuatro con cincuenta y seis: el punto separa miles y la coma decimales,
    al revés que en Python.
    """
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))

    s = str(valor).strip()
    for simbolo in ("$", "ARS", "U$S", "USD", " "):
        s = s.replace(simbolo, "")
    if "," in s:
        # Coma presente: es el separador decimal, los puntos son de miles.
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _a_entero(valor: Any) -> int | None:
    d = _a_decimal(valor)
    if d is None:
        return None
    try:
        return int(d)
    except (InvalidOperation, ValueError):
        return None


def detectar_encabezados(filas: list[list[Any]]) -> tuple[int, list[str]]:
    """Encuentra la fila de encabezados y devuelve (índice, encabezados).

    Heurística: la primera fila con al menos dos celdas de texto no vacías
    que coincidan con algún alias conocido. Buscar coincidencias y no solo
    "fila con varias celdas llenas" evita quedarse con el domicilio del
    proveedor, que también ocupa varias celdas.
    """
    for i, fila in enumerate(filas[:_MAX_FILAS_ENCABEZADO]):
        normalizadas = [normalizar(c) for c in fila]
        aciertos = sum(
            1 for celda in normalizadas
            if celda and any(celda in alias or celda.startswith(alias) for grupo in _ALIAS.values() for alias in grupo)
        )
        if aciertos >= 2:
            return i, [str(c).strip() if c is not None else "" for c in fila]

    raise ExcelInvalido(
        "No se encontró una fila de encabezados reconocible en las primeras "
        f"{_MAX_FILAS_ENCABEZADO} filas. Revisá que el archivo tenga una fila con "
        "los títulos de las columnas (código, descripción, cantidad, precio)."
    )


def sugerir_mapeo(encabezados: list[str]) -> dict[str, str]:
    """Propone un mapeo a partir de los encabezados, para pre-completar el
    formulario la primera vez que se importa de un proveedor.

    Es una sugerencia, no una decisión: el usuario la confirma o la corrige
    antes de que se guarde. Un mapeo mal adivinado y aplicado en silencio
    sería peor que no adivinar nada.
    """
    sugerencia: dict[str, str] = {}
    usados: set[str] = set()
    for campo, alias in _ALIAS.items():
        for encabezado in encabezados:
            if encabezado in usados:
                continue
            norm = normalizar(encabezado)
            if norm and any(norm == a or norm.startswith(a) for a in alias):
                sugerencia[campo] = encabezado
                usados.add(encabezado)
                break
    return sugerencia


def leer_filas(contenido: bytes) -> tuple[list[str], list[list[Any]], int]:
    """Abre el archivo y devuelve (encabezados, filas de datos, fila del encabezado).

    `read_only` y `data_only`: el primero evita cargar la planilla entera en
    memoria, el segundo devuelve el resultado de las fórmulas en vez de su
    texto -- una factura con `=C2*D2` en la columna de importe es común.
    """
    from io import BytesIO

    try:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl tira excepciones muy variadas
        raise ExcelInvalido(
            "No se pudo abrir el archivo. ¿Es un .xlsx válido? Los .xls viejos "
            "hay que guardarlos como .xlsx primero."
        ) from exc

    hoja = wb[wb.sheetnames[0]]
    filas = [list(f) for f in hoja.iter_rows(values_only=True)]
    wb.close()

    if not filas:
        raise ExcelInvalido("El archivo está vacío.")

    indice, encabezados = detectar_encabezados(filas)
    return encabezados, filas[indice + 1:], indice


def parsear(contenido: bytes, mapeo: dict[str, str]) -> list[FilaCruda]:
    """Convierte el Excel en filas usando el mapeo indicado.

    Las filas que no se pueden interpretar NO se descartan: se devuelven con
    `error` cargado. Descartarlas en silencio sería la peor opción posible --
    la factura tendría menos ítems de los reales y el total no cerraría sin
    que se sepa por qué.
    """
    encabezados, filas, offset = leer_filas(contenido)
    return mapear_filas(encabezados, filas, offset, mapeo)


def mapear_filas(
    encabezados: list[str], filas: list[list[Any]], offset: int, mapeo: dict[str, str]
) -> list[FilaCruda]:
    """Aplica un mapeo de columnas a filas ya extraídas.

    Vive separado de `parsear()` porque lo usan dos orígenes distintos: el
    Excel y las tablas que `pdf_import.py` saca de un PDF. Que compartan esta
    función es lo que garantiza que un "1.234,56" o una cantidad negativa se
    interpreten igual vengan de donde vengan -- si cada origen tuviera su
    propia lógica de coerción, tarde o temprano divergirían.
    """
    faltantes = [c for c in ("sku", "name", "quantity") if c not in mapeo]
    if faltantes:
        raise ExcelInvalido(f"Faltan columnas obligatorias en el mapeo: {', '.join(faltantes)}.")

    indices: dict[str, int] = {}
    for campo, encabezado in mapeo.items():
        objetivo = normalizar(encabezado)
        for i, actual in enumerate(encabezados):
            if normalizar(actual) == objetivo:
                indices[campo] = i
                break
        else:
            raise ExcelInvalido(
                f"La columna «{encabezado}» del mapeo guardado no está en este archivo. "
                "Puede que el proveedor haya cambiado el formato: revisá el mapeo."
            )

    def celda(fila: list[Any], campo: str) -> Any:
        i = indices.get(campo)
        return fila[i] if i is not None and i < len(fila) else None

    resultado: list[FilaCruda] = []
    for n, fila in enumerate(filas, start=offset + 2):  # +2: 1-based y saltando el encabezado
        sku = str(celda(fila, "sku") or "").strip()
        nombre = str(celda(fila, "name") or "").strip()
        if not sku and not nombre:
            continue  # fila vacía: separador o final de la tabla

        cantidad = _a_entero(celda(fila, "quantity"))
        costo = _a_decimal(celda(fila, "unit_cost"))

        error = None
        if not sku:
            error = "Sin código de artículo"
        elif len(sku) > 40:
            error = "El código supera los 40 caracteres"
        elif not nombre:
            error = "Sin descripción"
        elif cantidad is None:
            error = "Cantidad ilegible"
        elif cantidad <= 0:
            # Las notas de crédito vienen con cantidades negativas. No se
            # soportan todavía: se marcan para que el usuario decida, en vez
            # de restar stock sin querer.
            error = "Cantidad cero o negativa (¿nota de crédito?)"

        resultado.append(FilaCruda(
            row_number=n,
            sku=sku[:40],
            name=nombre[:160] or "(sin descripción)",
            quantity=cantidad if cantidad and cantidad > 0 else 0,
            unit_cost=costo,
            error=error,
        ))

    if not resultado:
        raise ExcelInvalido("No se encontró ninguna fila con datos debajo de los encabezados.")

    return resultado
